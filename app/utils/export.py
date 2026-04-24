import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _header_style():
    # синий фон с белым жирным текстом для строки заголовка таблицы
    fill = PatternFill('solid', fgColor='2563EB')
    font = Font(bold=True, color='FFFFFF')
    align = Alignment(horizontal='center', vertical='center')
    return fill, font, align


def _auto_width(ws, n_rows, n_cols, start_row=1):
    # подбираем ширину каждой колонки под самое длинное значение, но не больше 40 символов
    for col in range(1, n_cols + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or ''))
             for r in range(start_row, start_row + n_rows + 1)),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 40)


def export_leads_excel(leads, campaigns, single_campaign_name=None):
    """
    single_campaign_name задан → один лист «Лиды» с подзаголовком кампании, без листа кампаний.
    Иначе → два листа: «Лиды» и «Кампании» (только переданные кампании).
    ID лида вынесен в последний столбец.
    """
    wb = openpyxl.Workbook()
    fill, font, align = _header_style()

    # ── Лист 1: Лиды ──────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Лиды'

    lead_headers = [
        'Имя', 'Фамилия', 'Email', 'Телефон', 'Должность',
        'Компания', 'ИНН', 'Город', 'Тип клиента',
        'ЛПР (ФИО)', 'ЛПР (Должность)',
        'Статус', 'Score', 'Источник',
        'UTM Source', 'UTM Medium', 'UTM Campaign',
        'Кампания', 'Ответственный',
        'Сумма сделки', 'Заметки',
        'Создан', 'Конвертирован', 'ID',
    ]

    header_row = 1
    if single_campaign_name:
        title_cell = ws.cell(row=1, column=1, value=f'Кампания: {single_campaign_name}')
        title_cell.font = Font(bold=True, size=12)
        header_row = 2

    for col, h in enumerate(lead_headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = align

    ws.freeze_panes = f'A{header_row + 1}'
    ws.row_dimensions[header_row].height = 22

    for i, lead in enumerate(leads):
        r = header_row + 1 + i
        ws.cell(row=r, column=1,  value=lead.first_name)
        ws.cell(row=r, column=2,  value=lead.last_name)
        ws.cell(row=r, column=3,  value=lead.email)
        ws.cell(row=r, column=4,  value=lead.phone)
        ws.cell(row=r, column=5,  value=lead.position)
        ws.cell(row=r, column=6,  value=lead.company)
        ws.cell(row=r, column=7,  value=lead.inn)
        ws.cell(row=r, column=8,  value=lead.city)
        ws.cell(row=r, column=9,  value=lead.client_type.upper() if lead.client_type else None)
        ws.cell(row=r, column=10, value=lead.decision_maker_name)
        ws.cell(row=r, column=11, value=lead.decision_maker_position)
        ws.cell(row=r, column=12, value=lead.status)
        ws.cell(row=r, column=13, value=lead.score)
        ws.cell(row=r, column=14, value=lead.source)
        ws.cell(row=r, column=15, value=lead.utm_source)
        ws.cell(row=r, column=16, value=lead.utm_medium)
        ws.cell(row=r, column=17, value=lead.utm_campaign)
        ws.cell(row=r, column=18, value=lead.campaign.name if lead.campaign else None)
        ws.cell(row=r, column=19,
                value=(lead.assignee.full_name or lead.assignee.username)
                       if lead.assignee else None)
        ws.cell(row=r, column=20,
                value=float(lead.deal_amount) if lead.deal_amount else None)
        ws.cell(row=r, column=21, value=lead.notes)
        ws.cell(row=r, column=22,
                value=lead.created_at.strftime('%d.%m.%Y %H:%M')
                       if lead.created_at else None)
        ws.cell(row=r, column=23,
                value=lead.converted_at.strftime('%d.%m.%Y')
                       if lead.converted_at else None)
        ws.cell(row=r, column=24, value=lead.id)

    _auto_width(ws, len(leads), len(lead_headers), start_row=header_row)

    # ── Лист 2: Кампании (только для нескольких или всех) ─────────
    if not single_campaign_name:
        ws2 = wb.create_sheet('Кампании')

        camp_headers = [
            'Название', 'Статус', 'Канал', 'Аудитория',
            'Бюджет', 'Потрачено', 'Лидов',
            'Конверсия %', 'CPL', 'Выручка', 'ROI %',
            'Дата начала', 'Дата конца', 'ID',
        ]

        for col, h in enumerate(camp_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = font
            cell.fill = fill
            cell.alignment = align

        ws2.freeze_panes = 'A2'
        ws2.row_dimensions[1].height = 22

        for row, c in enumerate(campaigns, 2):
            ws2.cell(row=row, column=1,  value=c.name)
            ws2.cell(row=row, column=2,  value=c.status)
            ws2.cell(row=row, column=3,  value=c.channel)
            ws2.cell(row=row, column=4,  value=c.target_audience)
            ws2.cell(row=row, column=5,  value=float(c.budget) if c.budget else 0)
            ws2.cell(row=row, column=6,  value=float(c.spent) if c.spent else 0)
            ws2.cell(row=row, column=7,  value=c.lead_count)
            ws2.cell(row=row, column=8,  value=c.conversion_rate)
            ws2.cell(row=row, column=9,  value=c.cpl)
            ws2.cell(row=row, column=10, value=float(c.revenue) if c.revenue else 0)
            ws2.cell(row=row, column=11, value=c.roi)
            ws2.cell(row=row, column=12,
                    value=c.start_date.strftime('%d.%m.%Y') if c.start_date else None)
            ws2.cell(row=row, column=13,
                    value=c.end_date.strftime('%d.%m.%Y') if c.end_date else None)
            ws2.cell(row=row, column=14, value=c.id)

        _auto_width(ws2, len(campaigns), len(camp_headers))

    # сохраняем книгу в буфер в памяти и возвращаем его для отправки через send_file
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
